"""Core label generation logic."""

import csv
import html as html_module
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, TypedDict

from openpyxl import load_workbook

from ar_book_labels.layout import Layout

# ==================== AR Level Standard Colors ====================
LEVEL_COLORS: List[Tuple[float, float, str]] = [
    (0.1, 1.5, "#FFD700"),   # yellow
    (1.6, 2.0, "#2E8B57"),   # green
    (2.1, 2.5, "#00008B"),   # dark blue
    (2.6, 3.0, "#DC143C"),   # red
    (3.1, 3.5, "#FF69B4"),   # pink
    (3.6, 4.0, "#800080"),   # purple
    (4.1, 4.5, "#FF8C00"),   # orange
    (4.6, 5.0, "#00BFFF"),   # light blue
    (5.1, 5.5, "#FF6600"),   # neon orange
    (5.6, 6.0, "#39FF14"),   # neon green
    (6.1, 6.5, "#1C1C1C"),   # black
    (6.6, 99.0, "#8B4513"),  # brown
]

# Default column names (matching the template)
DEFAULT_COLUMNS: Dict[str, str] = {
    "title": "AR Title",
    "author": "AR Author",
    "level": "Book Level",
    "points": "AR Points",
    "quiz": "Quiz Number",
}
REQUIRED_FIELDS: List[str] = ["title", "author", "level", "points", "quiz"]


# ==================== TypedDict for Book Data ====================

class BookDict(TypedDict):
    """Represents a single book record extracted from the spreadsheet.

    Attributes
    ----------
    title:
        Book title.
    author:
        Book author.
    level:
        AR reading level (numeric).
    points:
        AR points value.
    quiz:
        Quiz number.
    """

    title: str
    author: str
    level: Any
    points: Any
    quiz: Any


def get_level_color(level: Any, colors: Optional[List[Tuple[float, float, str]]] = None) -> str:
    """Return the hex colour for a given AR level.

    Parameters
    ----------
    level:
        Numeric AR level (int, float, or string that can be cast to float).
    colors:
        Optional list of ``(low, high, hex)`` tuples.  When ``None`` the
        module-level ``LEVEL_COLORS`` constant is used.

    Returns
    -------
    str
        Hex colour string (e.g. ``"#FFD700"``), or ``"#999999"`` if the
        level is out of range or not numeric.
    """
    colors_to_use = colors if colors is not None else LEVEL_COLORS
    try:
        level = float(level)
    except (TypeError, ValueError):
        return "#999999"
    for low, high, color in colors_to_use:
        if low <= level <= high:
            return color
    return "#999999"


def get_badge_text_color(bg_color: str) -> str:
    """Return black or white text colour for readability on *bg_color*."""
    light = {"#FFD700", "#39FF14"}
    return "#000000" if bg_color in light else "#FFFFFF"


def split_text_lines(text: str, max_chars_per_line: int) -> List[str]:
    """Wrap text into at most 2 lines; truncate with ellipsis only if >2 lines."""
    text = str(text).strip()
    if not text:
        return [text]
    if len(text) <= max_chars_per_line:
        return [text]
    split_pos = text.rfind(" ", 0, max_chars_per_line)
    if split_pos == -1:
        split_pos = max_chars_per_line
    line1 = text[:split_pos].rstrip()
    remainder = text[split_pos:].lstrip()
    if len(remainder) <= max_chars_per_line:
        return [line1, remainder]
    return [line1, remainder[:max_chars_per_line - 1].rstrip() + "\u2026"]


def read_books(
    file_path: str,
    sheet_name: Optional[str] = None,
    columns: Optional[Dict[str, str]] = None,
    start_row: int = 2,
) -> Tuple[List[BookDict], List[str]]:
    """Read book data from an Excel (.xlsx) or CSV (.csv) file.

    The file format is auto-detected from the file extension.

    Parameters
    ----------
    file_path:
        Path to the .xlsx or .csv file.
    sheet_name:
        Name of the sheet to read (Excel only).  ``None`` means the first sheet.
        Ignored for CSV files.
    columns:
        Dict mapping internal keys to column names.
        Missing keys fall back to ``DEFAULT_COLUMNS``.
    start_row:
        1-indexed row number where data begins (1 = header row).

    Returns
    -------
    tuple[list[BookDict], list[str]]
        ``(books_list, warnings_list)``

    Raises
    ------
    ValueError
        If required columns are not found or the file extension is unsupported.
    FileNotFoundError
        If the file does not exist.
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _read_books_csv(file_path, columns=columns, start_row=start_row)
    elif suffix == ".xlsx":
        return _read_books_excel(
            file_path, sheet_name=sheet_name, columns=columns, start_row=start_row
        )
    else:
        raise ValueError(
            f"Unsupported file format: '{suffix}'. Use .xlsx or .csv."
        )


def _read_books_excel(
    excel_path: str,
    sheet_name: Optional[str],
    columns: Optional[Dict[str, str]],
    start_row: int,
) -> Tuple[List[BookDict], List[str]]:
    """Read book data from an Excel (.xlsx) file.

    Parameters
    ----------
    excel_path:
        Path to the .xlsx file.
    sheet_name:
        Name of the sheet to read.  ``None`` means the first sheet.
    columns:
        Dict mapping internal keys to Excel column names.
        Missing keys fall back to ``DEFAULT_COLUMNS``.
    start_row:
        1-indexed row number where data begins (1 = header row).

    Returns
    -------
    tuple[list[BookDict], list[str]]
        ``(books_list, warnings_list)``
    """
    cols = {**DEFAULT_COLUMNS, **(columns or {})}
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    if sheet_name is None:
        ws = wb.worksheets[0]
    else:
        ws = wb[sheet_name]
    actual_sheet_name = ws.title
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h) for h in rows[0] if h is not None]

    # Check that required columns exist in the sheet
    missing_cols = []
    for key in REQUIRED_FIELDS:
        if cols[key] not in headers:
            missing_cols.append(f"{key} ('{cols[key]}')")
    if missing_cols:
        wb.close()
        raise ValueError(
            f"Columns not found in sheet '{actual_sheet_name}': {', '.join(missing_cols)}.\n"
            f"Available columns: {', '.join(headers)}\n"
            f"Use --col-* options to map your column names."
        )

    books: List[BookDict] = []
    warnings: List[str] = []
    data_rows = rows[start_row - 1:]  # start_row is 1-indexed
    for i, row in enumerate(data_rows):
        row_num = start_row + i
        vals = list(row[:len(headers)])
        if all(v is None for v in vals):
            continue
        raw = dict(zip(headers, vals))

        # Validate required fields
        missing = []
        for key in REQUIRED_FIELDS:
            val = raw.get(cols[key])
            if val is None or (isinstance(val, str) and not val.strip()):
                missing.append(cols[key])
        if missing:
            warnings.append(f"Row {row_num}: skipped — missing: {', '.join(missing)}")
            continue

        books.append({
            "title": raw[cols["title"]],
            "author": raw[cols["author"]],
            "level": raw[cols["level"]],
            "points": raw[cols["points"]],
            "quiz": raw[cols["quiz"]],
        })

    wb.close()
    return books, warnings


def _read_books_csv(
    csv_path: str,
    columns: Optional[Dict[str, str]],
    start_row: int,
) -> Tuple[List[BookDict], List[str]]:
    """Read book data from a CSV file.

    Parameters
    ----------
    csv_path:
        Path to the .csv file.
    columns:
        Dict mapping internal keys to CSV column names.
        Missing keys fall back to ``DEFAULT_COLUMNS``.
    start_row:
        1-indexed row number where data begins (1 = header row).

    Returns
    -------
    tuple[list[BookDict], list[str]]
        ``(books_list, warnings_list)``

    Raises
    ------
    ValueError
        If required columns are not found in the CSV header row.
    """
    cols = {**DEFAULT_COLUMNS, **(columns or {})}

    with open(csv_path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        all_rows = list(reader)

    if not all_rows:
        raise ValueError(f"CSV file is empty: {csv_path}")

    headers = [h.strip() for h in all_rows[0] if h.strip()]

    # Check that required columns exist
    missing_cols = []
    for key in REQUIRED_FIELDS:
        if cols[key] not in headers:
            missing_cols.append(f"{key} ('{cols[key]}')")
    if missing_cols:
        raise ValueError(
            f"Columns not found in CSV file: {', '.join(missing_cols)}.\n"
            f"Available columns: {', '.join(headers)}\n"
            f"Use --col-* options to map your column names."
        )

    books: List[BookDict] = []
    warnings: List[str] = []
    data_rows = all_rows[start_row - 1:]  # start_row is 1-indexed

    for i, row in enumerate(data_rows):
        row_num = start_row + i
        # Pad row to header length if shorter
        padded_row = list(row) + [""] * (len(headers) - len(row))
        raw = dict(zip(headers, padded_row[:len(headers)]))

        # Skip entirely empty rows
        if all(v == "" or v is None for v in raw.values()):
            continue

        # Validate required fields
        missing = []
        for key in REQUIRED_FIELDS:
            val = raw.get(cols[key])
            if val is None or (isinstance(val, str) and not val.strip()):
                missing.append(cols[key])
        if missing:
            warnings.append(f"Row {row_num}: skipped — missing: {', '.join(missing)}")
            continue

        # Convert numeric fields where possible
        level_val = raw[cols["level"]]
        points_val = raw[cols["points"]]
        quiz_val = raw[cols["quiz"]]

        try:
            level_val = float(level_val)
        except (ValueError, TypeError):
            pass  # keep as string

        try:
            points_val = float(points_val)
        except (ValueError, TypeError):
            pass

        try:
            quiz_val = int(float(quiz_val))
        except (ValueError, TypeError):
            pass

        books.append({
            "title": raw[cols["title"]],
            "author": raw[cols["author"]],
            "level": level_val,
            "points": points_val,
            "quiz": quiz_val,
        })

    return books, warnings


def _generate_label_svg(
    book: BookDict,
    bw: bool = False,
    with_border: bool = False,
    layout: Optional[Layout] = None,
) -> str:
    """Generate SVG markup for a single label.

    Parameters
    ----------
    book:
        Book record with title, author, level, points, quiz.
    bw:
        When True, render in black-and-white mode — white circle with a
        thin black outline and a black level number.  When False (default),
        use the standard AR color-coded badge.
    with_border:
        When True, add a thin printable border around the label
        for manual cutting guides.
    layout:
        Optional :class:`Layout` instance.  When ``None`` a default
        50×30 mm layout is used (backward-compatible behaviour).

    Returns
    -------
    str
        SVG markup string for the label (a ``<g>`` element).
    """
    if layout is None:
        layout = Layout()

    title_raw = book.get("title", "") or ""
    level = book.get("level", 0)
    points = str(book.get("points", ""))
    quiz = str(book.get("quiz", ""))

    level_str = f"{float(level):.1f}" if isinstance(level, (int, float)) else str(level)
    badge_color = get_level_color(level, colors=layout.level_colors)
    badge_text = get_badge_text_color(badge_color)

    # bw mode overrides the badge fill/outline and the level number color.
    circle_fill = "white" if bw else badge_color
    circle_stroke = ' stroke="black" stroke-width="0.3"' if bw else ""
    level_fill = "black" if bw else badge_text

    author_raw = str(book.get("author", "") or "").strip()
    author_display = html_module.escape(
        author_raw if len(author_raw) <= 17 else author_raw[:16].rstrip() + "\u2026"
    )
    title_lines = [html_module.escape(ln) for ln in split_text_lines(title_raw, 14)]
    title_start = layout.top_y + layout.author_lh + 0.5
    title_ys = [title_start + i * layout.title_lh for i in range(len(title_lines))]

    # Determine max chars per line based on available width
    available_text_w = layout.label_w - layout.rx - 3  # 3mm right padding
    max_chars = max(10, int(available_text_w / 2.0))  # rough estimate

    p: List[str] = []
    p.append(
        f'<rect class="label-outline" x="0" y="0" '
        f'width="{layout.label_w}" height="{layout.label_h}" '
        f'rx="{layout.label_rx}" fill="none"/>'
    )
    if with_border:
        p.append(
            f'<rect class="label-border" x="0" y="0" '
            f'width="{layout.label_w}" height="{layout.label_h}" '
            f'rx="{layout.label_rx}" fill="none"/>'
        )
    p.append(
        f'<circle cx="{layout.cx}" cy="{layout.cy}" r="{layout.cr}" '
        f'fill="{circle_fill}"{circle_stroke}/>'
    )
    p.append(
        f'<text x="{layout.cx}" y="{layout.top_y}" text-anchor="middle" '
        f'dominant-baseline="hanging" fill="black" font-family="{layout.font_family}" '
        f'font-size="4.5" font-weight="700" letter-spacing="0.3">ATOS</text>'
    )
    p.append(
        f'<text x="{layout.cx}" y="{layout.cy + 2.5}" text-anchor="middle" '
        f'fill="{level_fill}" font-family="{layout.font_family}" '
        f'font-size="5.5" font-weight="700">{level_str}</text>'
    )
    p.append(
        f'<text x="{layout.rx}" y="{layout.top_y}" dominant-baseline="hanging" '
        f'fill="#555" font-family="{layout.font_family}" '
        f'font-size="2.6">{author_display}</text>'
    )
    for idx, y in enumerate(title_ys):
        p.append(
            f'<text x="{layout.rx}" y="{y}" dominant-baseline="hanging" '
            f'fill="black" font-family="{layout.font_family}" '
            f'font-size="3.2" font-weight="600">{title_lines[idx]}</text>'
        )
    p.append(
        f'<text x="{layout.rx}" y="{layout.points_y}" dominant-baseline="hanging" '
        f'fill="#666" font-family="{layout.font_family}" '
        f'font-size="2.5">Points:</text>'
    )
    p.append(
        f'<text x="{layout.vx}" y="{layout.points_y}" text-anchor="middle" '
        f'dominant-baseline="hanging" fill="black" font-family="{layout.font_family}" '
        f'font-size="2.8" font-weight="700">{points}</text>'
    )
    p.append(
        f'<text x="{layout.rx}" y="{layout.quiz_y}" dominant-baseline="hanging" '
        f'fill="#666" font-family="{layout.font_family}" '
        f'font-size="2.5">Quiz:</text>'
    )
    p.append(
        f'<text x="{layout.vx}" y="{layout.quiz_y}" text-anchor="middle" '
        f'dominant-baseline="hanging" fill="black" font-family="{layout.font_family}" '
        f'font-size="2.8" font-weight="700">{quiz}</text>'
    )

    return "<g>\n  " + "\n  ".join(p) + "\n</g>"


def build_html(
    books: List[BookDict],
    display_scale: int = 1,
    bw: bool = False,
    with_border: bool = False,
    layout: Optional[Layout] = None,
) -> str:
    """Build a multi-page HTML document with SVG labels.

    Page dimensions are in millimetres (A4 portrait: 210mm x 297mm) so that
    printing is 1:1.  On screen, each page is enlarged via CSS ``transform:
    scale(display_scale)``; ``@media print`` resets the transform to guarantee
    exact physical dimensions.

    Parameters
    ----------
    books:
        List of book dicts.
    display_scale:
        Scale factor for screen preview.
    bw:
        When True, render labels in black-and-white mode.
    with_border:
        When True, add a thin printable cutting-guide border
        around each label.
    layout:
        Optional :class:`Layout` instance.  When ``None`` a default
        50×30 mm layout is used (backward-compatible behaviour).

    Returns
    -------
    str
        Complete HTML document string.
    """
    if layout is None:
        layout = Layout()

    pages: List[List[BookDict]] = []
    for i in range(0, len(books), layout.labels_per_page):
        pages.append(books[i:i + layout.labels_per_page])

    s = display_scale

    page_blocks: List[str] = []
    for page_idx, page_books in enumerate(pages):
        labels_svg = ""
        for i, book in enumerate(page_books):
            row = i // layout.cols
            col = i % layout.cols
            x = layout.cols_x[col]
            y = layout.rows_y[row]
            labels_svg += (
                f'<g transform="translate({x},{y})">'
                f'{_generate_label_svg(book, bw=bw, with_border=with_border, layout=layout)}'
                f'</g>\n'
            )

        pb = ' style="page-break-after: always;"' if page_idx < len(pages) - 1 else ""
        page_blocks.append(
            f'<div class="page"{pb}>\n'
        f'<svg width="{layout.page_w:g}mm" height="{layout.page_h:g}mm" '
        f'viewBox="0 0 {layout.page_w:g} {layout.page_h:g}" '
        f'xmlns="http://www.w3.org/2000/svg">\n'
        f'<rect width="{layout.page_w:g}" height="{layout.page_h:g}" fill="white"/>\n'
            f'{labels_svg}'
            f'</svg>\n'
            f'</div>'
        )

    return (
        '<!DOCTYPE html>\n'
        '<html lang="en">\n'
        '<head>\n'
        '<meta charset="UTF-8">\n'
        '<title>AR Book Labels</title>\n'
        '<style>\n'
        f'  @page {{ size: {layout.page_w:g}mm {layout.page_h:g}mm; margin: 0; }}\n'
        '  * { margin: 0; padding: 0; box-sizing: border-box; }\n'
        '  body {\n'
        '    background: #e0e0e0;\n'
        '    -webkit-print-color-adjust: exact !important;\n'
        '    print-color-adjust: exact !important;\n'
        '    color-adjust: exact !important;\n'
        '  }\n'
        '  .page {\n'
        f'    width: {layout.page_w:g}mm;\n'
        f'    height: {layout.page_h:g}mm;\n'
        '    margin: 0 auto;\n'
        f'    margin-bottom: calc(10px + {layout.page_h:g}mm * ({s} - 1));\n'
        '    background: white;\n'
        f'    transform: scale({s});\n'
        '    transform-origin: top center;\n'
        '  }\n'
        '  .page svg { display: block; }\n'
        '  .label-outline {\n'
        '    fill: none;\n'
        '    stroke: #999999;\n'
        '    stroke-width: 0.2;\n'
        '    stroke-dasharray: 1.2, 1.2;\n'
        '  }\n'
        '  .label-border {\n'
        '    fill: none;\n'
        '    stroke: #000000;\n'
        '    stroke-width: 0.15;\n'
        '  }\n'
        '  @media print {\n'
        '    body { margin: 0; background: white; }\n'
        '    .page {\n'
        '      transform: none;\n'
        '      margin: 0 !important;\n'
        '      page-break-after: always;\n'
        '    }\n'
        '    .page:last-child { page-break-after: auto; }\n'
        '    .label-outline { stroke: none; }\n'
        '  }\n'
        '</style>\n'
        '</head>\n'
        '<body>\n'
        f'{"".join(page_blocks)}\n'
        '</body>\n'
        '</html>'
    )


def generate(
    file_path: str,
    output_path: str,
    sheet_name: Optional[str] = None,
    column_mapping: Optional[Dict[str, str]] = None,
    start_row: int = 2,
    display_scale: int = 1,
    bw: bool = False,
    with_border: bool = False,
    layout: Optional[Layout] = None,
) -> Tuple[int, int, List[str]]:
    """High-level entry point: read Excel/CSV, generate labels, write HTML.

    Parameters
    ----------
    file_path:
        Path to the .xlsx or .csv file.
    output_path:
        Output HTML file path.
    sheet_name:
        Sheet name to read (Excel only); ``None`` means the first sheet.
    column_mapping:
        Optional dict mapping internal keys to column names.
    start_row:
        1-indexed row number where data begins.
    display_scale:
        Scale factor for screen preview.
    bw:
        When True, render labels in black-and-white mode (white circle,
        thin black outline, black level number).
    with_border:
        When True, add a thin printable cutting-guide border
        around each label.
    layout:
        Optional :class:`Layout` instance.  When ``None`` a default
        50×30 mm layout is used (backward-compatible behaviour).

    Returns
    -------
    tuple[int, int, list[str]]
        ``(number_of_books, number_of_pages, warnings_list)``
    """
    if layout is None:
        layout = Layout()

    books, warnings = read_books(
        file_path, sheet_name, columns=column_mapping, start_row=start_row
    )

    # T01: Auto-create parent directories for the output path
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    html_content = build_html(books, display_scale, bw=bw, with_border=with_border, layout=layout)

    with open(out, "w", encoding="utf-8") as f:
        f.write(html_content)

    n_pages = 0 if not books else (len(books) + layout.labels_per_page - 1) // layout.labels_per_page
    return len(books), n_pages, warnings
