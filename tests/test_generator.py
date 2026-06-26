"""Tests for ar-book-labels generator."""

import csv
from pathlib import Path

import pytest
from openpyxl import Workbook

from ar_book_labels.generator import (
    get_level_color,
    get_badge_text_color,
    split_text_lines,
    read_books,
    build_html,
    generate,
    BookDict,
    LEVEL_COLORS,
    DEFAULT_COLUMNS,
    REQUIRED_FIELDS,
)
from ar_book_labels.layout import Layout


# ---------------------------------------------------------------------------
# Original tests (backward compatibility)
# ---------------------------------------------------------------------------

def test_level_colors():
    assert get_level_color(1.0) == "#FFD700"  # yellow
    assert get_level_color(1.8) == "#2E8B57"  # green
    assert get_level_color(2.3) == "#00008B"  # dark blue
    assert get_level_color(2.8) == "#DC143C"  # red
    assert get_level_color(3.3) == "#FF69B4"  # pink
    assert get_level_color(3.8) == "#800080"  # purple
    assert get_level_color(4.3) == "#FF8C00"  # orange
    assert get_level_color(4.8) == "#00BFFF"  # light blue
    assert get_level_color(5.1) == "#FF6600"  # neon orange
    assert get_level_color(5.8) == "#39FF14"  # neon green
    assert get_level_color(6.3) == "#1C1C1C"  # black
    assert get_level_color(7.0) == "#8B4513"  # brown
    assert get_level_color(0.0) == "#999999"  # out of range → grey
    assert get_level_color("invalid") == "#999999"


def test_level_colors_custom():
    """get_level_color with a custom colour list."""
    custom = [(0.0, 3.0, "#111111"), (3.1, 99.0, "#EEEEEE")]
    assert get_level_color(2.0, colors=custom) == "#111111"
    assert get_level_color(5.0, colors=custom) == "#EEEEEE"
    assert get_level_color(0.0, colors=custom) == "#111111"


def test_badge_text_color():
    assert get_badge_text_color("#FFD700") == "#000000"  # yellow → black text
    assert get_badge_text_color("#39FF14") == "#000000"  # neon green → black text
    assert get_badge_text_color("#00008B") == "#FFFFFF"  # dark blue → white text
    assert get_badge_text_color("#1C1C1C") == "#FFFFFF"  # black → white text


def test_split_text_lines_single():
    assert split_text_lines("Short", 19) == ["Short"]
    assert split_text_lines("", 19) == [""]


def test_split_text_lines_wrap():
    result = split_text_lines("Diary of a Wimpy Kid: The Last Straw", 19)
    assert len(result) == 2
    assert result[0] == "Diary of a Wimpy"
    assert result[1] == "Kid: The Last Straw"


def test_split_text_lines_truncate():
    # This text needs 3 lines at 19 chars, so line 2 gets truncated
    result = split_text_lines("The Journey to the Center of the Earth is Great", 19)
    assert len(result) == 2
    assert result[0] == "The Journey to the"
    assert result[1].endswith("\u2026")  # ellipsis on truncated line 2


def test_split_text_lines_no_space():
    # Very long single word exceeding 2 lines
    result = split_text_lines("A" * 50, 19)
    assert len(result) == 2
    assert result[1].endswith("\u2026")


def test_build_html_empty():
    html = build_html([])
    assert "<!DOCTYPE html>" in html
    assert "AR Book Labels" in html
    assert "<svg" not in html  # no pages


def test_build_html_single_book():
    book = {
        "title": "Test Book",
        "author": "Test Author",
        "level": 3.5,
        "points": 5,
        "quiz": 12345,
    }
    html = build_html([book])
    assert "Test Book" in html
    assert "Test Author" in html
    assert "12345" in html
    assert html.count('<div class="page"') == 1


def test_build_html_multi_page():
    default_layout = Layout()
    books = [
        {
            "title": f"Book {i}",
            "author": f"Author {i}",
            "level": 2.0,
            "points": 1,
            "quiz": i,
        }
        for i in range(default_layout.labels_per_page + 1)
    ]
    html = build_html(books)
    assert html.count('<div class="page"') == 2


# ---------------------------------------------------------------------------
# New tests: Layout-based API
# ---------------------------------------------------------------------------

def test_build_html_with_custom_layout():
    """build_html with a custom Layout should produce correct dimensions."""
    layout = Layout.from_config({"label_size": "70x37"})
    book = {
        "title": "Test Book",
        "author": "Test Author",
        "level": 3.5,
        "points": 5,
        "quiz": 12345,
    }
    html = build_html([book], layout=layout)
    # Page dimensions should match the layout
    assert 'width="210mm"' in html
    assert 'height="297mm"' in html
    # Label should be present
    assert "Test Book" in html
    assert "Test Author" in html


def test_build_html_with_custom_colors():
    """Custom level_colors should be used when provided."""
    custom_colors = [(0.0, 99.0, "#FF0000")]  # all red
    layout = Layout(level_colors=custom_colors)
    book = {
        "title": "Red Book",
        "author": "Author",
        "level": 3.0,
        "points": 1,
        "quiz": 100,
    }
    html = build_html([book], layout=layout)
    # The badge should use the custom colour
    assert '#FF0000' in html


def test_build_html_with_custom_font():
    """Custom font_family should appear in the SVG."""
    layout = Layout(font_family="Arial, sans-serif")
    book = {
        "title": "Font Test",
        "author": "Author",
        "level": 2.0,
        "points": 1,
        "quiz": 100,
    }
    html = build_html([book], layout=layout)
    assert "Arial, sans-serif" in html


def test_build_html_default_layout_backward_compat():
    """build_html() without layout should produce the same output as before."""
    book = {
        "title": "Compat Test",
        "author": "Author",
        "level": 2.0,
        "points": 1,
        "quiz": 100,
    }
    # Call without layout (backward compatible)
    html_no_layout = build_html([book])
    # Call with default Layout
    html_default_layout = build_html([book], layout=Layout())
    # They should be identical
    assert html_no_layout == html_default_layout


def test_build_html_page_count_with_custom_layout():
    """Page count should respect layout.labels_per_page."""
    layout = Layout.from_config({"label_size": "70x37"})  # 21 labels/page
    books = [
        {"title": f"Book {i}", "author": "A", "level": 2.0, "points": 1, "quiz": i}
        for i in range(22)  # 22 books → 2 pages (21 + 1)
    ]
    html = build_html(books, layout=layout)
    assert html.count('<div class="page"') == 2


def test_build_html_bw_mode_with_layout():
    """B&W mode should work correctly with a custom layout."""
    layout = Layout(label_w=70, label_h=37, margin_x=0, margin_y=0, col_gap=0, row_gap=0)
    book = {
        "title": "BW Test",
        "author": "Author",
        "level": 3.0,
        "points": 1,
        "quiz": 100,
    }
    html = build_html([book], bw=True, layout=layout)
    assert 'fill="white"' in html  # circle fill in bw mode
    assert 'stroke="black"' in html  # circle stroke in bw mode


# ---------------------------------------------------------------------------
# Helper: create a test Excel file
# ---------------------------------------------------------------------------

def _create_test_excel(
    tmp_path: Path,
    rows: list,
    headers: list = None,
    sheet_name: str = "Sheet1",
    filename: str = "test.xlsx",
) -> Path:
    """Create a temporary Excel file for testing.

    Parameters
    ----------
    tmp_path:
        pytest tmp_path fixture.
    rows:
        List of dicts, each dict representing a data row.
    headers:
        List of header strings.  Defaults to the standard AR columns.
    sheet_name:
        Name for the worksheet.
    filename:
        Filename within tmp_path.

    Returns
    -------
    Path
        Path to the created .xlsx file.
    """
    if headers is None:
        headers = ["AR Title", "AR Author", "Book Level", "AR Points", "Quiz Number"]

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for row_data in rows:
        ws.append([row_data.get(h) for h in headers])
    dest = tmp_path / filename
    wb.save(dest)
    wb.close()
    return dest


def _create_test_csv(
    tmp_path: Path,
    rows: list,
    headers: list = None,
    filename: str = "test.csv",
) -> Path:
    """Create a temporary CSV file for testing.

    Parameters
    ----------
    tmp_path:
        pytest tmp_path fixture.
    rows:
        List of dicts, each dict representing a data row.
    headers:
        List of header strings.  Defaults to the standard AR columns.
    filename:
        Filename within tmp_path.

    Returns
    -------
    Path
        Path to the created .csv file.
    """
    if headers is None:
        headers = ["AR Title", "AR Author", "Book Level", "AR Points", "Quiz Number"]

    dest = tmp_path / filename
    with open(dest, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        for row_data in rows:
            writer.writerow(row_data)
    return dest


# ---------------------------------------------------------------------------
# T03: generate() function tests
# ---------------------------------------------------------------------------

class TestGenerateFunction:
    """Comprehensive tests for the generate() entry point."""

    def test_normal_generate(self, tmp_path: Path):
        """Normal generation flow with a valid Excel file."""
        excel = _create_test_excel(tmp_path, [
            {"AR Title": "Book A", "AR Author": "Author A", "Book Level": 3.5, "AR Points": 5, "Quiz Number": 1001},
            {"AR Title": "Book B", "AR Author": "Author B", "Book Level": 2.0, "AR Points": 3, "Quiz Number": 1002},
        ])
        output = tmp_path / "out" / "labels.html"
        n_books, n_pages, warnings = generate(str(excel), str(output))
        assert n_books == 2
        assert n_pages == 1
        assert warnings == []
        assert output.exists()
        content = output.read_text(encoding="utf-8")
        assert "Book A" in content
        assert "Book B" in content

    def test_empty_excel(self, tmp_path: Path):
        """Empty Excel file (only headers) should produce 0 books."""
        excel = _create_test_excel(tmp_path, rows=[])
        output = tmp_path / "labels.html"
        n_books, n_pages, warnings = generate(str(excel), str(output))
        assert n_books == 0
        assert n_pages == 0
        assert output.exists()
        content = output.read_text(encoding="utf-8")
        assert "<!DOCTYPE html>" in content

    def test_bw_mode(self, tmp_path: Path):
        """generate() in black-and-white mode should pass bw=True through."""
        excel = _create_test_excel(tmp_path, [
            {"AR Title": "BW Book", "AR Author": "Author", "Book Level": 3.0, "AR Points": 1, "Quiz Number": 100},
        ])
        output = tmp_path / "bw.html"
        generate(str(excel), str(output), bw=True)
        content = output.read_text(encoding="utf-8")
        assert 'fill="white"' in content
        assert 'stroke="black"' in content

    def test_with_border(self, tmp_path: Path):
        """generate() with with_border=True should add cutting-guide borders."""
        excel = _create_test_excel(tmp_path, [
            {"AR Title": "Border Book", "AR Author": "Author", "Book Level": 2.5, "AR Points": 2, "Quiz Number": 200},
        ])
        output = tmp_path / "border.html"
        generate(str(excel), str(output), with_border=True)
        content = output.read_text(encoding="utf-8")
        assert 'class="label-border"' in content

    def test_custom_layout(self, tmp_path: Path):
        """generate() should pass the layout to build_html correctly."""
        excel = _create_test_excel(tmp_path, [
            {"AR Title": "Custom Book", "AR Author": "Author", "Book Level": 4.0, "AR Points": 8, "Quiz Number": 300},
        ])
        output = tmp_path / "custom.html"
        layout = Layout.from_config({"label_size": "70x37"})
        n_books, n_pages, _ = generate(str(excel), str(output), layout=layout)
        assert n_books == 1
        content = output.read_text(encoding="utf-8")
        assert "Custom Book" in content

    def test_output_directory_created(self, tmp_path: Path):
        """generate() should auto-create parent directories for output."""
        excel = _create_test_excel(tmp_path, [
            {"AR Title": "Dir Book", "AR Author": "Author", "Book Level": 1.0, "AR Points": 1, "Quiz Number": 1},
        ])
        output = tmp_path / "deep" / "nested" / "dir" / "labels.html"
        generate(str(excel), str(output))
        assert output.exists()

    def test_warnings_for_incomplete_rows(self, tmp_path: Path):
        """Rows with missing required fields should generate warnings."""
        excel = _create_test_excel(tmp_path, [
            {"AR Title": "Good Book", "AR Author": "Author", "Book Level": 2.0, "AR Points": 3, "Quiz Number": 100},
            {"AR Title": "Bad Book", "AR Author": "", "Book Level": None, "AR Points": 1, "Quiz Number": 200},
        ])
        output = tmp_path / "warn.html"
        n_books, _, warnings = generate(str(excel), str(output))
        assert n_books == 1  # only the good row
        assert len(warnings) == 1
        assert "Row 3" in warnings[0]

    def test_custom_column_mapping(self, tmp_path: Path):
        """generate() should honour a custom column mapping."""
        headers = ["Title", "Writer", "Lvl", "Pts", "Quiz#"]
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        ws.append(["My Book", "My Writer", 4.5, 10, 9999])
        excel = tmp_path / "custom_cols.xlsx"
        wb.save(excel)
        wb.close()

        output = tmp_path / "out.html"
        column_mapping = {
            "title": "Title",
            "author": "Writer",
            "level": "Lvl",
            "points": "Pts",
            "quiz": "Quiz#",
        }
        n_books, _, _ = generate(str(excel), str(output), column_mapping=column_mapping)
        assert n_books == 1
        content = output.read_text(encoding="utf-8")
        assert "My Book" in content


# ---------------------------------------------------------------------------
# T03: Error path tests
# ---------------------------------------------------------------------------

class TestGenerateErrors:
    """Tests for error paths in generate/read_books."""

    def test_invalid_excel_file(self, tmp_path: Path):
        """Non-Excel file should raise an error."""
        bad_file = tmp_path / "bad.xlsx"
        bad_file.write_text("not an excel file", encoding="utf-8")
        output = tmp_path / "out.html"
        with pytest.raises(Exception):
            generate(str(bad_file), str(output))

    def test_missing_required_column(self, tmp_path: Path):
        """Excel file missing a required column should raise ValueError."""
        wb = Workbook()
        ws = wb.active
        ws.append(["AR Title", "AR Author"])  # missing Level, Points, Quiz
        ws.append(["Book", "Author"])
        excel = tmp_path / "missing_cols.xlsx"
        wb.save(excel)
        wb.close()
        output = tmp_path / "out.html"
        with pytest.raises(ValueError, match="Columns not found"):
            generate(str(excel), str(output))

    def test_invalid_sheet_name(self, tmp_path: Path):
        """Requesting a non-existent sheet should raise KeyError."""
        excel = _create_test_excel(tmp_path, [
            {"AR Title": "Book", "AR Author": "Author", "Book Level": 2.0, "AR Points": 1, "Quiz Number": 1},
        ])
        output = tmp_path / "out.html"
        with pytest.raises(KeyError):
            generate(str(excel), str(output), sheet_name="NonExistentSheet")


# ---------------------------------------------------------------------------
# T07: TypedDict tests
# ---------------------------------------------------------------------------

class TestBookDict:
    """Verify BookDict TypedDict is accessible and typed correctly."""

    def test_book_dict_is_accessible(self):
        """BookDict should be importable from generator module."""
        assert BookDict is not None

    def test_book_dict_annotation_in_read_books(self, tmp_path: Path):
        """read_books should return a list of BookDicts."""
        excel = _create_test_excel(tmp_path, [
            {"AR Title": "Book", "AR Author": "Author", "Book Level": 2.0, "AR Points": 1, "Quiz Number": 1},
        ])
        books, _ = read_books(str(excel), None)
        assert len(books) == 1
        assert set(books[0].keys()) == {"title", "author", "level", "points", "quiz"}


# ---------------------------------------------------------------------------
# Existing constants should still work
# ---------------------------------------------------------------------------

class TestModuleConstants:
    """Verify that non-deprecated module constants are still directly accessible."""

    def test_level_colors_accessible(self):
        assert len(LEVEL_COLORS) == 12
        assert LEVEL_COLORS[0] == (0.1, 1.5, "#FFD700")

    def test_default_columns_accessible(self):
        assert DEFAULT_COLUMNS["title"] == "AR Title"
        assert DEFAULT_COLUMNS["author"] == "AR Author"

    def test_required_fields_accessible(self):
        assert REQUIRED_FIELDS == ["title", "author", "level", "points", "quiz"]


# ---------------------------------------------------------------------------
# T05: CSV input tests
# ---------------------------------------------------------------------------

class TestCSVInput:
    """Tests for CSV file input support."""

    def test_read_books_csv_basic(self, tmp_path: Path):
        """read_books should read a basic CSV file."""
        csv_file = _create_test_csv(tmp_path, [
            {"AR Title": "CSV Book 1", "AR Author": "Author 1", "Book Level": "2.5", "AR Points": "3", "Quiz Number": "100"},
            {"AR Title": "CSV Book 2", "AR Author": "Author 2", "Book Level": "4.0", "AR Points": "5", "Quiz Number": "200"},
        ])
        books, warnings = read_books(str(csv_file), None)
        assert len(books) == 2
        assert len(warnings) == 0
        assert books[0]["title"] == "CSV Book 1"
        assert books[1]["title"] == "CSV Book 2"

    def test_read_books_csv_numeric_conversion(self, tmp_path: Path):
        """CSV reader should convert numeric fields."""
        csv_file = _create_test_csv(tmp_path, [
            {"AR Title": "Book", "AR Author": "Author", "Book Level": "3.5", "AR Points": "7", "Quiz Number": "1234"},
        ])
        books, _ = read_books(str(csv_file), None)
        assert len(books) == 1
        assert books[0]["level"] == 3.5
        assert books[0]["points"] == 7.0
        assert books[0]["quiz"] == 1234

    def test_read_books_csv_missing_columns(self, tmp_path: Path):
        """CSV with missing required columns should raise ValueError."""
        csv_file = _create_test_csv(tmp_path, [
            {"Title": "Book", "Author": "Author"},
        ], headers=["Title", "Author"])
        with pytest.raises(ValueError, match="Columns not found"):
            read_books(str(csv_file), None)

    def test_read_books_csv_empty_file(self, tmp_path: Path):
        """Empty CSV file should raise ValueError."""
        csv_file = tmp_path / "empty.csv"
        csv_file.write_text("", encoding="utf-8")
        with pytest.raises(ValueError, match="empty"):
            read_books(str(csv_file), None)

    def test_read_books_csv_skip_incomplete_rows(self, tmp_path: Path):
        """CSV rows with missing required fields should be skipped with warning."""
        csv_file = _create_test_csv(tmp_path, [
            {"AR Title": "Good Book", "AR Author": "Author", "Book Level": "2.0", "AR Points": "3", "Quiz Number": "100"},
            {"AR Title": "Bad Book", "AR Author": "", "Book Level": "", "AR Points": "1", "Quiz Number": "200"},
        ])
        books, warnings = read_books(str(csv_file), None)
        assert len(books) == 1
        assert len(warnings) == 1
        assert "Row 3" in warnings[0]

    def test_read_books_csv_custom_columns(self, tmp_path: Path):
        """CSV with custom column names should work with column mapping."""
        headers = ["Title", "Writer", "Lvl", "Pts", "Quiz#"]
        csv_file = _create_test_csv(tmp_path, [
            {"Title": "My Book", "Writer": "My Writer", "Lvl": "4.5", "Pts": "10", "Quiz#": "9999"},
        ], headers=headers)
        column_mapping = {
            "title": "Title",
            "author": "Writer",
            "level": "Lvl",
            "points": "Pts",
            "quiz": "Quiz#",
        }
        books, _ = read_books(str(csv_file), None, columns=column_mapping)
        assert len(books) == 1
        assert books[0]["title"] == "My Book"
        assert books[0]["author"] == "My Writer"

    def test_read_books_csv_utf8_bom(self, tmp_path: Path):
        """CSV with UTF-8 BOM should be handled correctly."""
        csv_file = tmp_path / "bom.csv"
        content = "\ufeffAR Title,AR Author,Book Level,AR Points,Quiz Number\nBOM Book,Author,2.0,3,100"
        csv_file.write_text(content, encoding="utf-8")
        books, _ = read_books(str(csv_file), None)
        assert len(books) == 1
        assert books[0]["title"] == "BOM Book"

    def test_generate_from_csv(self, tmp_path: Path):
        """generate() should work with CSV input files."""
        csv_file = _create_test_csv(tmp_path, [
            {"AR Title": "CSV Generate", "AR Author": "Author", "Book Level": "3.0", "AR Points": "5", "Quiz Number": "500"},
        ])
        output = tmp_path / "output.html"
        n_books, n_pages, warnings = generate(str(csv_file), str(output))
        assert n_books == 1
        assert n_pages == 1
        assert output.exists()
        content = output.read_text(encoding="utf-8")
        assert "CSV Generate" in content

    def test_read_books_unsupported_format(self, tmp_path: Path):
        """Unsupported file format should raise ValueError."""
        txt_file = tmp_path / "test.txt"
        txt_file.write_text("some text", encoding="utf-8")
        with pytest.raises(ValueError, match="Unsupported file format"):
            read_books(str(txt_file), None)
